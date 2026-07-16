"""Recent-slice reader over the NICE cancer-recommendations spreadsheet.

The Storyblok xlsx is the *index* of cancer TAs: TA id + fiscal year + technology +
indication + categorisation, at fiscal-year granularity. This turns the recent slice
(fiscal_year_start >= cutoff.year - 2) into typed ``NiceTaRef``s - the work-list the
per-TA guidance fetcher resolves to exact dates.

Column detection is value-based where names are unreliable (the "TA ID" header is missed
by name matching, so the TA column is found by the ``TA\\d+`` value regex). Pure +
unit-tested against an in-memory workbook fixture; no network.
"""
from __future__ import annotations

import io
import re

from pydantic import BaseModel

from strata_platform.sources.nice import classify_recommendation, detect_rec_col, looks_like_xlsx
from strata_platform.sources.nice_feasibility import detect_year_col, parse_fiscal_year

_TA_RE = re.compile(r"^\s*TA\d+\s*$", re.IGNORECASE)
_MIN_COL_HITS = 5

# Documented column layout (1-indexed in the spec -> 0-indexed here), used only as a
# fallback when header-name detection fails: TA(0) Year(1) Technology(3) Indication(5).
_FALLBACK = {"technology": 3, "indication": 5}


class NiceTaRef(BaseModel):
    """One row of the cancer index, recent slice."""

    model_config = {"frozen": True}

    ta_id: str
    fiscal_year_start: int
    categorisation: str  # classified outcome label (see sources.nice classifier)
    technology: str
    indication: str


def _cell(row: tuple, j: int | None) -> str:
    if j is None or j >= len(row) or row[j] is None:
        return ""
    return str(row[j]).strip()


def detect_ta_col(body: list[tuple]) -> int | None:
    """Column whose values look like TA ids. The header 'TA ID' is missed by name
    matching, so detect by content."""
    ncols = max((len(r) for r in body), default=0)
    if not ncols:
        return None
    scores = [
        sum(1 for r in body if j < len(r) and _TA_RE.match(str(r[j] or "")))
        for j in range(ncols)
    ]
    j = max(range(ncols), key=scores.__getitem__)
    return j if scores[j] >= _MIN_COL_HITS else None


def _name_col(header: list[str], *names: str) -> int | None:
    for j, h in enumerate(header):
        if any(n in h for n in names):
            return j
    return None


def recent_cancer_tas(xbytes: bytes, cutoff) -> list[NiceTaRef]:
    """Parse the recent-slice (FY >= cutoff.year - 2) cancer TAs into NiceTaRefs."""
    import openpyxl

    if not looks_like_xlsx(xbytes):
        raise RuntimeError("NICE index is not a valid xlsx (HTML/redirect?)")
    wb = openpyxl.load_workbook(io.BytesIO(xbytes), read_only=True, data_only=True)
    ws = next(
        (wb[s] for s in wb.sheetnames if "recommend" in s.lower()),
        wb[wb.sheetnames[0]],
    )
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).lower() if c is not None else "" for c in rows[0]]
    body = [r for r in rows[1:] if r and not all(c is None for c in r)]

    c_ta = detect_ta_col(body)
    c_year = detect_year_col(header, body)
    c_cat = detect_rec_col(body)
    if c_ta is None:
        raise RuntimeError("could not detect the TA-id column in the NICE index")
    if c_year is None:
        raise RuntimeError("could not detect the fiscal-year column in the NICE index")
    c_tech = _name_col(header, "technology") or _FALLBACK["technology"]
    c_ind = _name_col(header, "indication") or _FALLBACK["indication"]

    min_year = cutoff.year - 2
    refs: list[NiceTaRef] = []
    for r in body:
        ta = _cell(r, c_ta)
        if not _TA_RE.match(ta):
            continue
        fy = parse_fiscal_year(r[c_year]) if c_year < len(r) else None
        if fy is None or fy < min_year:
            continue
        refs.append(
            NiceTaRef(
                ta_id=ta.upper(),
                fiscal_year_start=fy,
                categorisation=classify_recommendation(_cell(r, c_cat)),
                technology=_cell(r, c_tech),
                indication=_cell(r, c_ind),
            )
        )
    return refs
