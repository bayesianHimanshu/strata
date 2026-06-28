"""STRATA — NICE feasibility summary (fiscal-year based).

FINDING (Phase 0): the NICE cancer-recommendations spreadsheet has NO decision
date — its only temporal field is "Year of Publication" as a UK fiscal year
("YYYY/YY", e.g. 2025/26 = Apr 2025–Mar 2026). The recommendation lives in
"Categorisation (for specific recommendation)". So this sheet is the INDEX of
cancer TAs (universe + categorisation + rough year), not the decision record.

Consequence: fiscal-year granularity cannot cleanly resolve a mid-year model
cutoff. With cutoff 2026-02-01, only FY 2026/27+ is cleanly post-cutoff; FY
2025/26 straddles it and needs day-resolution to split. Exact post-cutoff dates
(and the committee rationale for the gold set) come from each per-TA guidance page
(nice.org.uk/guidance/taXXXX → "Published: DD Month YYYY" + the rationale).

MERGE: add parse_fiscal_year / fy_bucket / detect_year_col / nice_feasibility into
sources/nice.py. Also add the 'only_in_research' pattern to _REC_PATTERNS (NICE's
"Only in research"/OIR categorisation) and to NEGATIVE_OUTCOMES.
"""
from __future__ import annotations

import io
import re
from collections import Counter
from datetime import date

from sources.nice import (  # reuse the canonical helpers already in the module
    classify_recommendation,
    detect_rec_col,
    looks_like_xlsx,
)

_FY = re.compile(r"^(\d{4})/(\d{2})$")
RESTRICTED = frozenset(
    {"not_recommended", "optimised", "non_submission", "only_in_research"}
)


def parse_fiscal_year(v) -> int | None:
    """'2025/26' -> 2025 (start year; FY runs Apr(start)..Mar(start+1)). None if not
    a fiscal year. The '/YY' must be the last two digits of start+1."""
    if v is None:
        return None
    m = _FY.match(str(v).strip())
    if not m:
        return None
    start = int(m.group(1))
    return start if (start + 1) % 100 == int(m.group(2)) else None


def fy_bucket(start_year: int, cutoff: date) -> str:
    """Where a fiscal year sits relative to the model cutoff.
    'post' = unambiguously after; 'straddle' = contains the cutoff (needs per-TA
    dates to split); 'pre' = before."""
    if date(start_year + 1, 3, 31) < cutoff:
        return "pre"
    if date(start_year, 4, 1) > cutoff:
        return "post"
    return "straddle"


def detect_year_col(header: list[str], body: list[tuple]) -> int | None:
    for j, h in enumerate(header):
        if "year" in h:
            return j
    ncols = max((len(r) for r in body), default=0)
    if not ncols:
        return None
    scores = [
        sum(1 for r in body if j < len(r) and parse_fiscal_year(r[j]) is not None)
        for j in range(ncols)
    ]
    j = max(range(ncols), key=scores.__getitem__)
    return j if scores[j] >= 5 else None


def nice_feasibility(xbytes: bytes, cutoff: date) -> dict:
    """Year-based feasibility summary over the NICE cancer-recommendations sheet."""
    import openpyxl

    if not looks_like_xlsx(xbytes):
        raise RuntimeError("NICE download is not a valid xlsx (HTML/redirect?)")
    wb = openpyxl.load_workbook(io.BytesIO(xbytes), read_only=True, data_only=True)
    ws = next(
        (wb[s] for s in wb.sheetnames if "recommend" in s.lower()),
        wb[wb.sheetnames[0]],
    )
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"error": "empty NICE workbook"}
    header = [str(c).lower() if c is not None else "" for c in rows[0]]
    body = [r for r in rows[1:] if r and not all(c is None for c in r)]

    c_year = detect_year_col(header, body)
    c_rec = detect_rec_col(body)
    if c_year is None:
        raise RuntimeError("no fiscal-year column found in NICE sheet")

    by_year: Counter = Counter()
    by_bucket: Counter = Counter()
    categorisation: Counter = Counter()
    restricted_recent: Counter = Counter()
    for r in body:
        y = parse_fiscal_year(r[c_year]) if c_year < len(r) else None
        out = (
            classify_recommendation(str(r[c_rec]))
            if c_rec is not None and c_rec < len(r) and r[c_rec]
            else "other"
        )
        categorisation[out] += 1
        if y is None:
            continue
        by_year[y] += 1
        b = fy_bucket(y, cutoff)
        by_bucket[b] += 1
        if b in ("straddle", "post") and out in RESTRICTED:
            restricted_recent[b] += 1

    return {
        "total_rows": len(body),
        "columns": {"fiscal_year": c_year, "recommendation": c_rec},
        "by_bucket": dict(by_bucket),
        "recent_fiscal_years": dict(
            sorted((y, c) for y, c in by_year.items() if y >= cutoff.year - 2)
        ),
        "categorisation_breakdown": dict(categorisation),
        "restricted_recent": dict(restricted_recent),
        "note": (
            "fiscal-year granularity; 'post' is cleanly post-cutoff, 'straddle' "
            "(FY containing the cutoff) needs per-TA guidance dates to split"
        ),
    }


def nice_verdict(nice: dict, *, clean_min: int = 12) -> dict:
    """Verdict from the year-based summary. The cleanly-post restricted count is the
    only NICE-spreadsheet-native clean-arm signal; the straddle restricted count is
    recoverable once per-TA dates are fetched."""
    if "error" in nice or "restricted_recent" not in nice:
        return {"arm_a_clean": "unknown", "reason": "NICE summary failed"}
    post = nice["restricted_recent"].get("post", 0)
    straddle = nice["restricted_recent"].get("straddle", 0)
    clean_ok = post >= clean_min
    return {
        "clean_min": clean_min,
        "post_cutoff_restricted_clean": post,
        "straddle_restricted_recoverable": straddle,
        "arm_a_clean_self_sufficient": clean_ok,
        "recommendation": (
            "Proceed; NICE post-cutoff slice alone powers the clean Arm A arm."
            if clean_ok else
            f"NICE spreadsheet alone yields only {post} cleanly-post restricted "
            f"decisions, but ~{straddle} more sit in the straddle FY and are "
            "recoverable by fetching per-TA guidance dates. Resolve those (bounded "
            "fetch of the recent slice) and/or pull G-BA/IQWiG forward; the per-TA "
            "fetch also yields the committee rationale needed for the gold set."
        ),
    }