"""NICE cancer-appraisal recommendations client (primary gold source).

No clean JSON API: locate the cancer-recommendations spreadsheet on the page and
parse it into Decision rows. Written defensively because NICE rearranges URLs and
columns. The xlsx-link finder, the recommendation classifier, the column
detectors, and the workbook parser are all pure functions, unit-tested without the
network.

This module is the SINGLE parser for the NICE workbook. The Phase 0 probe must
call parse_workbook_detailed here rather than reimplementing the loop — duplicating
it is how the column-detection fix silently failed to take effect.
"""
from __future__ import annotations

import io
import re
import time
from datetime import date

import httpx

from core.config import (
    CLEAN_ARM_MIN_NEGATIVE,
    MODEL_CUTOFF,
    NICE_BASE,
    NICE_CANCER_PAGE,
)
from core.contracts import Decision
from core.provenance import SourceRecord, normalize_date, snapshot
from sources.base import build_client

# Order matters: most specific / most negative first so 'not recommended' is not
# swallowed by the generic 'recommend' pattern.
_REC_PATTERNS: list[tuple[str, str]] = [
    ("non_submission", r"non[- ]?submission|terminated"),
    ("not_recommended", r"not recommend"),
    ("only_in_research", r"only in research|\boir\b"),
    ("cdf", r"cancer drugs fund|\bcdf\b"),
    ("optimised", r"optimis|restrict"),
    ("recommended", r"recommend"),
]

# Outcomes that signal a restricted / negative committee position — the slice that
# carries the most evidence-gap signal for Arm A.
NEGATIVE_OUTCOMES = frozenset(
    {"not_recommended", "optimised", "non_submission", "only_in_research"}
)

# A detected column must have at least this many positive hits to be accepted.
# Absolute, not relative to row count: the NICE sheet is dominated by older rows,
# so a relative floor can reject the true (sparser) date column. The correct date
# column wins argmax over every other column regardless of sparsity.
MIN_COL_HITS = 5


def classify_recommendation(text: str) -> str:
    t = (text or "").lower()
    for label, pat in _REC_PATTERNS:
        if re.search(pat, t):
            return label
    return "other"


def looks_like_xlsx(b: bytes) -> bool:
    """xlsx/xlsm are zip containers — guard against HTML/redirect pages being fed
    to openpyxl (the InvalidFileException failure mode)."""
    return b[:4] == b"PK\x03\x04"


def find_xlsx_url(html: str, base: str = NICE_CANCER_PAGE) -> str | None:
    for m in re.finditer(r'href="([^"]+\.xlsx[^"]*)"', html, re.IGNORECASE):
        href = m.group(1)
        if href.startswith("http"):
            return href
        if href.startswith("/"):
            return NICE_BASE + href
        return base.rsplit("/", 1)[0] + "/" + href
    return None


# --------------------------------------------------------------------------- #
# Column detection — value-based, not header-name-based.
#
# Header-name matching collided (a header like "Date recommendation issued"
# matches both the date and recommendation patterns, resolving them to the same
# column). Detect by what the cells ACTUALLY contain: the date column is the one
# whose values parse as dates; the recommendation column is the one whose values
# classify to a known outcome. The two can never collide — different predicates.
# --------------------------------------------------------------------------- #

def _argmax_col(body: list[tuple], scorer) -> int | None:
    if not body:
        return None
    ncols = max((len(r) for r in body), default=0)
    if ncols == 0:
        return None
    scores = [sum(1 for r in body if j < len(r) and scorer(r[j])) for j in range(ncols)]
    j = max(range(ncols), key=scores.__getitem__)
    return j if scores[j] >= MIN_COL_HITS else None


def detect_date_col(body: list[tuple]) -> int | None:
    return _argmax_col(body, lambda v: normalize_date(v) is not None)


def detect_rec_col(body: list[tuple]) -> int | None:
    return _argmax_col(
        body, lambda v: bool(v) and classify_recommendation(str(v)) != "other"
    )


def parse_workbook_detailed(xbytes: bytes) -> tuple[list[Decision], dict]:
    """Parse the workbook into (dated Decision rows, diagnostics).

    Diagnostics carry total_rows, undated count, detected column indices, and the
    recommendation breakdown over ALL non-empty rows — everything the Phase 0
    feasibility report needs, so the probe never has to reparse. Date and
    recommendation columns are detected by cell content, so column reordering or
    header-name overlap cannot misalign them.
    """
    import openpyxl  # local import: keep openpyxl off the hot import path

    if not looks_like_xlsx(xbytes):
        raise RuntimeError(
            "NICE download is not a valid xlsx (got HTML/redirect?) — the link "
            "finder matched a non-file URL or the page is gated"
        )

    wb = openpyxl.load_workbook(io.BytesIO(xbytes), read_only=True, data_only=True)
    ws = next(
        (wb[s] for s in wb.sheetnames if "recommend" in s.lower()),
        wb[wb.sheetnames[0]],
    )
    rows = list(ws.iter_rows(values_only=True))
    empty_diag = {
        "total_rows": 0, "undated": 0,
        "columns_detected": {"date": None, "recommendation": None,
                             "ta": None, "title": None},
        "recommendation_breakdown": {},
    }
    if not rows:
        return [], empty_diag

    # Header row located by name only to find where data begins and to locate the
    # non-critical TA/title columns. Date/recommendation are detected by value.
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        cells = [str(c).lower() if c is not None else "" for c in row]
        if any("date" in c for c in cells) and any(
            "recommend" in c or "ta" in c for c in cells
        ):
            header_idx = i
            break
    header = [str(c).lower() if c is not None else "" for c in rows[header_idx]]

    def col(*names: str) -> int | None:
        for j, h in enumerate(header):
            if any(n in h for n in names):
                return j
        return None

    c_ta = col("ta number", "ta no", "appraisal", "reference")
    c_title = col("title", "technology", "name")

    body = [
        row
        for row in rows[header_idx + 1 :]
        if row is not None and not all(c is None for c in row)
    ]
    c_date = detect_date_col(body)
    c_rec = detect_rec_col(body)
    if c_date is None:
        raise RuntimeError(
            "could not detect a date column in the NICE workbook — inspect the "
            "sheet; leakage stratification cannot proceed without dates"
        )

    decisions: list[Decision] = []
    undated = 0
    breakdown: dict[str, int] = {}
    for n, row in enumerate(body):
        rtext = (
            str(row[c_rec])
            if c_rec is not None and c_rec < len(row) and row[c_rec]
            else ""
        )
        outcome = classify_recommendation(rtext)
        breakdown[outcome] = breakdown.get(outcome, 0) + 1

        d: date | None = normalize_date(row[c_date]) if c_date < len(row) else None
        if d is None:
            undated += 1
            continue

        title = (
            str(row[c_title])
            if c_title is not None and c_title < len(row) and row[c_title]
            else ""
        )
        ta = (
            str(row[c_ta])
            if c_ta is not None and c_ta < len(row) and row[c_ta]
            else f"NICE-row-{header_idx + 1 + n}"
        )
        decisions.append(
            Decision(
                agency="NICE",
                decision_id=ta,
                decision_date=d,
                indication=title,
                outcome=outcome,
                rationale_raw="",
            )
        )
    
    if len(body) and undated > 0.5 * len(body):
        raise RuntimeError(
            f"{undated}/{len(body)} NICE rows undated — date column (idx {c_date}) "
            f"misdetected or dates not parsing; check normalize_date / cell format"
        )

    diag = {
        "total_rows": len(body),
        "undated": undated,
        "columns_detected": {"date": c_date, "recommendation": c_rec,
                             "ta": c_ta, "title": c_title},
        "recommendation_breakdown": breakdown,
    }
    return decisions, diag


def parse_workbook(xbytes: bytes) -> list[Decision]:
    """Dated Decision rows only. Thin wrapper over parse_workbook_detailed."""
    return parse_workbook_detailed(xbytes)[0]


# Per-TA guidance fetching + parsing lives in sources/nice_guidance.py (the single
# guidance parser). resolve_clean_arm stays here: it operates on assembled
# Decisions, not on raw pages.


def resolve_clean_arm(decisions: list[Decision], cutoff: date = MODEL_CUTOFF) -> dict:
    """Exact pre/post-cutoff stratification once per-TA day-resolution dates exist
    (invariant #4). Post-cutoff = decision_date strictly after the model cutoff —
    the leakage-clean Arm A test set; the restricted slice carries the gap signal."""
    post = [d for d in decisions if d.decision_date > cutoff]
    pre = [d for d in decisions if d.decision_date <= cutoff]
    post_restricted = [d for d in post if d.outcome in NEGATIVE_OUTCOMES]
    return {
        "model_cutoff": cutoff.isoformat(),
        "total": len(decisions),
        "pre_cutoff": len(pre),
        "post_cutoff": len(post),
        "post_cutoff_restricted": len(post_restricted),
        "post_cutoff_restricted_ids": [d.decision_id for d in post_restricted],
        "clean_arm_self_sufficient": len(post_restricted) >= CLEAN_ARM_MIN_NEGATIVE,
        "clean_arm_min_required": CLEAN_ARM_MIN_NEGATIVE,
    }


def _get_with_retry(
    client: httpx.Client,
    url: str,
    *,
    retries: int = 4,
    headers: dict[str, str] | None = None,
) -> httpx.Response:
    """GET with exponential backoff on 429 and 5xx. The NICE page 502'd in Phase 0
    while the asset host stayed up — so callers prefer the known asset URL and lean
    on this for the transient page fetches."""
    last: httpx.Response | None = None
    for attempt in range(retries):
        r = client.get(url, headers=headers)
        if r.status_code == 429 or 500 <= r.status_code < 600:
            last = r
            time.sleep(2**attempt)
            continue
        r.raise_for_status()
        return r
    assert last is not None
    last.raise_for_status()
    return last


class NICEClient:
    def __init__(self, client: httpx.Client | None = None) -> None:
        self._client = client or build_client()

    def fetch_recommendations(
        self, *, xlsx_url_override: str | None = None
    ) -> tuple[list[Decision], SourceRecord]:
        """Locate, snapshot, and parse the cancer-recommendations workbook. Prefer a
        known asset URL (override) over scraping the page, which has 502'd."""
        xlsx_url = xlsx_url_override
        if not xlsx_url:
            page = _get_with_retry(
                self._client, NICE_CANCER_PAGE, headers={"Accept": "text/html"}
            )
            xlsx_url = find_xlsx_url(page.text)
            if not xlsx_url:
                raise RuntimeError(
                    "could not locate NICE cancer-recommendations .xlsx link "
                    "(pass xlsx_url_override)"
                )
        resp = _get_with_retry(
            self._client, xlsx_url, headers={"Accept": "application/octet-stream"}
        )
        xbytes = resp.content
        if not looks_like_xlsx(xbytes):
            raise RuntimeError(
                f"NICE link {xlsx_url} did not return a valid xlsx "
                f"(content-type {resp.headers.get('content-type')!r})"
            )
        rec = snapshot(
            xbytes, source="nice", source_id="cancer_recommendations", url=xlsx_url
        )
        return parse_workbook(xbytes), rec